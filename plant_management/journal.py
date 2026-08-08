import json
import sqlite3
import time
import uuid
from datetime import date, datetime, timedelta
from io import BytesIO
from pathlib import Path

DB_FILE = Path(__file__).resolve().parent.parent / "data.db"
UPLOAD_DIR = Path(__file__).resolve().parent.parent / "instance" / "journal_uploads"

MAX_ATTACHMENT_BYTES = 8 * 1024 * 1024
MAX_ATTACHMENTS_PER_REQUEST = 5

JOURNAL_CATEGORIES = (
    ("observation", "Beobachtung", "👁️", "#38bdf8"),
    ("care", "Pflege / Arbeit", "✂️", "#22c55e"),
    ("irrigation", "Bewässerung", "💧", "#06b6d4"),
    ("nutrition", "Nährstoffe", "🧪", "#84cc16"),
    ("training", "Training", "🪢", "#a855f7"),
    ("measurement", "Messung", "📈", "#6366f1"),
    ("issue", "Abweichung / Problem", "⚠️", "#f97316"),
    ("treatment", "Maßnahme / Behandlung", "🛠️", "#ef4444"),
    ("harvest", "Ernte", "📦", "#f59e0b"),
    ("note", "Allgemeine Notiz", "📝", "#94a3b8"),
)

JOURNAL_CATEGORY_MAP = {
    code: {
        "code": code,
        "label": label,
        "icon": icon,
        "color": color,
    }
    for code, label, icon, color in JOURNAL_CATEGORIES
}

JOURNAL_SEVERITIES = (
    ("info", "Info"),
    ("attention", "Beobachten"),
    ("critical", "Kritisch"),
)

JOURNAL_SEVERITY_LABELS = dict(JOURNAL_SEVERITIES)

MEASUREMENT_FIELDS = (
    ("temperature", "Temperatur", "°C"),
    ("humidity", "Luftfeuchte", "%"),
    ("vpd", "VPD", "kPa"),
    ("ph", "pH", ""),
    ("ec", "EC", "mS/cm"),
    ("water_temperature", "Wassertemperatur", "°C"),
    ("water_amount", "Wassermenge", "L"),
    ("drain_amount", "Drain", "L"),
)

ALLOWED_ATTACHMENTS = {
    ".jpg": ("image/jpeg", b"\xff\xd8\xff"),
    ".jpeg": ("image/jpeg", b"\xff\xd8\xff"),
    ".png": ("image/png", b"\x89PNG\r\n\x1a\n"),
    ".webp": ("image/webp", b"RIFF"),
    ".pdf": ("application/pdf", b"%PDF"),
}


def _db():
    conn = sqlite3.connect(DB_FILE, check_same_thread=False)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA foreign_keys = ON")
    return conn


def _now():
    return int(time.time())


def _text(value):
    if value is None:
        return ""
    return str(value).strip()


def _normalize_datetime(value=None):
    if not value:
        return datetime.now().replace(second=0, microsecond=0).isoformat(timespec="minutes")

    text = str(value).strip()
    try:
        parsed = datetime.fromisoformat(text)
    except ValueError as exc:
        raise ValueError("Ungültiges Datum oder ungültige Uhrzeit.") from exc

    return parsed.replace(second=0, microsecond=0).isoformat(timespec="minutes")


def _normalize_date(value=None):
    if not value:
        return None
    text = str(value).strip()
    try:
        return date.fromisoformat(text[:10]).isoformat()
    except ValueError as exc:
        raise ValueError("Ungültiges Datum.") from exc


def _truthy(value):
    if isinstance(value, bool):
        return value
    if value is None:
        return False
    return str(value).strip().lower() in {"1", "true", "yes", "ja", "on", "x"}


def _to_float(value):
    if value is None or str(value).strip() == "":
        return None
    try:
        return float(str(value).replace(",", "."))
    except ValueError as exc:
        raise ValueError(f"Ungültiger Messwert: {value}") from exc


def init_plant_journal_db():
    UPLOAD_DIR.mkdir(parents=True, exist_ok=True)

    db = _db()
    try:
        db.executescript(
            """
            CREATE TABLE IF NOT EXISTS pm_journal_entries (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                occurred_at TEXT NOT NULL,
                category TEXT NOT NULL,
                severity TEXT NOT NULL DEFAULT 'info',
                title TEXT NOT NULL,
                body TEXT,
                tags TEXT,
                follow_up_required INTEGER NOT NULL DEFAULT 0,
                follow_up_due_on TEXT,
                resolved_at INTEGER,
                resolved_by INTEGER,
                resolved_by_name TEXT,
                source TEXT NOT NULL DEFAULT 'manual',
                created_by INTEGER,
                created_by_name TEXT,
                created_at INTEGER NOT NULL,
                updated_at INTEGER NOT NULL,
                cancelled_at INTEGER,
                cancelled_by INTEGER,
                cancelled_by_name TEXT,
                cancel_reason TEXT
            );

            CREATE INDEX IF NOT EXISTS idx_pm_journal_occurred
            ON pm_journal_entries(occurred_at DESC);

            CREATE INDEX IF NOT EXISTS idx_pm_journal_category
            ON pm_journal_entries(category);

            CREATE INDEX IF NOT EXISTS idx_pm_journal_followup
            ON pm_journal_entries(follow_up_required, resolved_at);

            CREATE TABLE IF NOT EXISTS pm_journal_entry_plants (
                entry_id INTEGER NOT NULL,
                plant_id INTEGER NOT NULL,
                PRIMARY KEY(entry_id, plant_id),
                FOREIGN KEY(entry_id) REFERENCES pm_journal_entries(id) ON DELETE CASCADE,
                FOREIGN KEY(plant_id) REFERENCES pm_plants(id) ON DELETE CASCADE
            );

            CREATE INDEX IF NOT EXISTS idx_pm_journal_entry_plants_plant
            ON pm_journal_entry_plants(plant_id, entry_id);

            CREATE TABLE IF NOT EXISTS pm_journal_entry_batches (
                entry_id INTEGER NOT NULL,
                batch_id INTEGER NOT NULL,
                PRIMARY KEY(entry_id, batch_id),
                FOREIGN KEY(entry_id) REFERENCES pm_journal_entries(id) ON DELETE CASCADE,
                FOREIGN KEY(batch_id) REFERENCES pm_batches(id) ON DELETE CASCADE
            );

            CREATE INDEX IF NOT EXISTS idx_pm_journal_entry_batches_batch
            ON pm_journal_entry_batches(batch_id, entry_id);

            CREATE TABLE IF NOT EXISTS pm_journal_measurements (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                entry_id INTEGER NOT NULL,
                metric TEXT NOT NULL,
                label TEXT NOT NULL,
                value REAL NOT NULL,
                unit TEXT,
                created_at INTEGER NOT NULL,
                FOREIGN KEY(entry_id) REFERENCES pm_journal_entries(id) ON DELETE CASCADE
            );

            CREATE INDEX IF NOT EXISTS idx_pm_journal_measurements_entry
            ON pm_journal_measurements(entry_id);

            CREATE TABLE IF NOT EXISTS pm_journal_attachments (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                entry_id INTEGER NOT NULL,
                stored_name TEXT NOT NULL UNIQUE,
                original_name TEXT NOT NULL,
                mime_type TEXT NOT NULL,
                size_bytes INTEGER NOT NULL,
                uploaded_at INTEGER NOT NULL,
                uploaded_by INTEGER,
                uploaded_by_name TEXT,
                deleted_at INTEGER,
                deleted_by INTEGER,
                deleted_by_name TEXT,
                FOREIGN KEY(entry_id) REFERENCES pm_journal_entries(id) ON DELETE CASCADE
            );

            CREATE INDEX IF NOT EXISTS idx_pm_journal_attachments_entry
            ON pm_journal_attachments(entry_id);

            CREATE TABLE IF NOT EXISTS pm_journal_revisions (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                entry_id INTEGER NOT NULL,
                revision_no INTEGER NOT NULL,
                snapshot_json TEXT NOT NULL,
                changed_at INTEGER NOT NULL,
                changed_by INTEGER,
                changed_by_name TEXT,
                FOREIGN KEY(entry_id) REFERENCES pm_journal_entries(id) ON DELETE CASCADE
            );

            CREATE UNIQUE INDEX IF NOT EXISTS idx_pm_journal_revision_unique
            ON pm_journal_revisions(entry_id, revision_no);
            """
        )
        db.commit()
    finally:
        db.close()


def _category_meta(code):
    return JOURNAL_CATEGORY_MAP.get(
        code,
        {
            "code": code,
            "label": code,
            "icon": "📝",
            "color": "#94a3b8",
        },
    )


def _decorate_entry(entry):
    item = dict(entry)
    meta = _category_meta(item["category"])
    item["category_label"] = meta["label"]
    item["category_icon"] = meta["icon"]
    item["category_color"] = meta["color"]
    item["severity_label"] = JOURNAL_SEVERITY_LABELS.get(
        item["severity"],
        item["severity"],
    )
    item["is_cancelled"] = bool(item.get("cancelled_at"))
    item["is_resolved"] = bool(item.get("resolved_at"))
    item["is_open_follow_up"] = bool(
        item.get("follow_up_required")
        and not item.get("resolved_at")
        and not item.get("cancelled_at")
    )
    item["occurred_date"] = str(item["occurred_at"])[:10]
    item["occurred_time"] = (
        str(item["occurred_at"])[11:16]
        if len(str(item["occurred_at"])) >= 16
        else ""
    )
    return item


def _load_links(db, entry_id):
    plants = [
        dict(row)
        for row in db.execute(
            """
            SELECT p.id, p.code, p.display_name, p.current_stage
            FROM pm_journal_entry_plants l
            JOIN pm_plants p ON p.id = l.plant_id
            WHERE l.entry_id = ?
            ORDER BY p.display_name COLLATE NOCASE
            """,
            (entry_id,),
        ).fetchall()
    ]

    batches = [
        dict(row)
        for row in db.execute(
            """
            SELECT b.id, b.code, b.name
            FROM pm_journal_entry_batches l
            JOIN pm_batches b ON b.id = l.batch_id
            WHERE l.entry_id = ?
            ORDER BY b.name COLLATE NOCASE
            """,
            (entry_id,),
        ).fetchall()
    ]

    measurements = [
        dict(row)
        for row in db.execute(
            """
            SELECT *
            FROM pm_journal_measurements
            WHERE entry_id = ?
            ORDER BY id
            """,
            (entry_id,),
        ).fetchall()
    ]

    attachments = [
        dict(row)
        for row in db.execute(
            """
            SELECT *
            FROM pm_journal_attachments
            WHERE entry_id = ? AND deleted_at IS NULL
            ORDER BY id
            """,
            (entry_id,),
        ).fetchall()
    ]

    return plants, batches, measurements, attachments


def _hydrate_entry(db, row):
    if not row:
        return None
    item = _decorate_entry(row)
    plants, batches, measurements, attachments = _load_links(db, item["id"])
    item["plants"] = plants
    item["batches"] = batches
    item["measurements"] = measurements
    item["attachments"] = attachments
    item["plant_ids"] = [p["id"] for p in plants]
    item["batch_ids"] = [b["id"] for b in batches]
    return item


def list_journal_entries(
    *,
    search=None,
    category=None,
    severity=None,
    plant_id=None,
    batch_id=None,
    date_from=None,
    date_to=None,
    open_follow_up=False,
    include_cancelled=False,
    limit=250,
):
    db = _db()
    try:
        where = []
        params = []

        if not include_cancelled:
            where.append("e.cancelled_at IS NULL")

        if search:
            q = f"%{str(search).strip()}%"
            where.append(
                "(e.title LIKE ? OR e.body LIKE ? OR e.tags LIKE ? "
                "OR e.created_by_name LIKE ?)"
            )
            params.extend([q, q, q, q])

        if category:
            where.append("e.category = ?")
            params.append(category)

        if severity:
            where.append("e.severity = ?")
            params.append(severity)

        if plant_id:
            where.append(
                """
                EXISTS (
                    SELECT 1
                    FROM pm_journal_entry_plants lp
                    WHERE lp.entry_id = e.id AND lp.plant_id = ?
                )
                """
            )
            params.append(int(plant_id))

        if batch_id:
            where.append(
                """
                EXISTS (
                    SELECT 1
                    FROM pm_journal_entry_batches lb
                    WHERE lb.entry_id = e.id AND lb.batch_id = ?
                )
                """
            )
            params.append(int(batch_id))

        if date_from:
            where.append("substr(e.occurred_at, 1, 10) >= ?")
            params.append(_normalize_date(date_from))

        if date_to:
            where.append("substr(e.occurred_at, 1, 10) <= ?")
            params.append(_normalize_date(date_to))

        if open_follow_up:
            where.append(
                "e.follow_up_required = 1 "
                "AND e.resolved_at IS NULL "
                "AND e.cancelled_at IS NULL"
            )

        sql = "SELECT e.* FROM pm_journal_entries e"
        if where:
            sql += " WHERE " + " AND ".join(where)
        sql += " ORDER BY e.occurred_at DESC, e.id DESC LIMIT ?"
        params.append(max(1, min(int(limit), 5000)))

        rows = db.execute(sql, params).fetchall()
        return [_hydrate_entry(db, row) for row in rows]
    finally:
        db.close()


def get_journal_entry(entry_id):
    db = _db()
    try:
        row = db.execute(
            "SELECT * FROM pm_journal_entries WHERE id = ?",
            (int(entry_id),),
        ).fetchone()
        return _hydrate_entry(db, row)
    finally:
        db.close()


def _parse_ids(values):
    result = []
    for value in values or []:
        try:
            parsed = int(value)
        except (TypeError, ValueError):
            continue
        if parsed > 0 and parsed not in result:
            result.append(parsed)
    return result


def measurements_from_form(data):
    measurements = []

    for metric, label, unit in MEASUREMENT_FIELDS:
        raw = data.get(f"measurement_{metric}")
        if raw is None or str(raw).strip() == "":
            continue

        measurements.append(
            {
                "metric": metric,
                "label": label,
                "value": _to_float(raw),
                "unit": unit,
            }
        )

    return measurements


def _entry_snapshot(db, entry_id):
    entry = _hydrate_entry(
        db,
        db.execute(
            "SELECT * FROM pm_journal_entries WHERE id = ?",
            (entry_id,),
        ).fetchone(),
    )
    if not entry:
        return None

    # Nur JSON-fähige Daten.
    return {
        "entry": {
            key: value
            for key, value in entry.items()
            if key not in {"plants", "batches", "measurements", "attachments"}
        },
        "plants": entry["plants"],
        "batches": entry["batches"],
        "measurements": entry["measurements"],
        "attachments": [
            {
                "id": item["id"],
                "original_name": item["original_name"],
                "mime_type": item["mime_type"],
                "size_bytes": item["size_bytes"],
            }
            for item in entry["attachments"]
        ],
    }


def _write_revision(db, entry_id, changed_by=None, changed_by_name=None):
    snapshot = _entry_snapshot(db, entry_id)
    if snapshot is None:
        return

    row = db.execute(
        """
        SELECT COALESCE(MAX(revision_no), 0) AS max_revision
        FROM pm_journal_revisions
        WHERE entry_id = ?
        """,
        (entry_id,),
    ).fetchone()

    revision_no = int(row["max_revision"]) + 1

    db.execute(
        """
        INSERT INTO pm_journal_revisions (
            entry_id, revision_no, snapshot_json,
            changed_at, changed_by, changed_by_name
        )
        VALUES (?, ?, ?, ?, ?, ?)
        """,
        (
            entry_id,
            revision_no,
            json.dumps(snapshot, ensure_ascii=False),
            _now(),
            changed_by,
            changed_by_name,
        ),
    )


def save_journal_entry(
    data,
    *,
    plant_ids=None,
    batch_ids=None,
    measurements=None,
    user_id=None,
    user_name=None,
    entry_id=None,
):
    category = _text(data.get("category")) or "note"
    severity = _text(data.get("severity")) or "info"

    if category not in JOURNAL_CATEGORY_MAP:
        raise ValueError("Ungültige Journal-Kategorie.")

    if severity not in JOURNAL_SEVERITY_LABELS:
        raise ValueError("Ungültige Priorität.")

    title = _text(data.get("title"))
    if not title:
        raise ValueError("Der Titel darf nicht leer sein.")

    body = _text(data.get("body")) or None
    tags = _text(data.get("tags")) or None
    occurred_at = _normalize_datetime(data.get("occurred_at"))

    follow_up_required = 1 if _truthy(data.get("follow_up_required")) else 0
    follow_up_due_on = (
        _normalize_date(data.get("follow_up_due_on"))
        if follow_up_required and data.get("follow_up_due_on")
        else None
    )

    plant_ids = _parse_ids(plant_ids)
    batch_ids = _parse_ids(batch_ids)
    measurements = measurements or []

    now = _now()
    db = _db()

    try:
        if entry_id:
            entry_id = int(entry_id)
            existing = db.execute(
                "SELECT id FROM pm_journal_entries WHERE id = ?",
                (entry_id,),
            ).fetchone()
            if not existing:
                raise ValueError("Journal-Eintrag wurde nicht gefunden.")

            _write_revision(
                db,
                entry_id,
                changed_by=user_id,
                changed_by_name=user_name,
            )

            db.execute(
                """
                UPDATE pm_journal_entries SET
                    occurred_at = ?,
                    category = ?,
                    severity = ?,
                    title = ?,
                    body = ?,
                    tags = ?,
                    follow_up_required = ?,
                    follow_up_due_on = ?,
                    updated_at = ?
                WHERE id = ?
                """,
                (
                    occurred_at,
                    category,
                    severity,
                    title,
                    body,
                    tags,
                    follow_up_required,
                    follow_up_due_on,
                    now,
                    entry_id,
                ),
            )

            db.execute(
                "DELETE FROM pm_journal_entry_plants WHERE entry_id = ?",
                (entry_id,),
            )
            db.execute(
                "DELETE FROM pm_journal_entry_batches WHERE entry_id = ?",
                (entry_id,),
            )
            db.execute(
                "DELETE FROM pm_journal_measurements WHERE entry_id = ?",
                (entry_id,),
            )
        else:
            cur = db.execute(
                """
                INSERT INTO pm_journal_entries (
                    occurred_at, category, severity, title, body, tags,
                    follow_up_required, follow_up_due_on,
                    source, created_by, created_by_name,
                    created_at, updated_at
                )
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, 'manual', ?, ?, ?, ?)
                """,
                (
                    occurred_at,
                    category,
                    severity,
                    title,
                    body,
                    tags,
                    follow_up_required,
                    follow_up_due_on,
                    user_id,
                    user_name,
                    now,
                    now,
                ),
            )
            entry_id = int(cur.lastrowid)

        for plant_id in plant_ids:
            db.execute(
                """
                INSERT OR IGNORE INTO pm_journal_entry_plants (
                    entry_id, plant_id
                )
                VALUES (?, ?)
                """,
                (entry_id, plant_id),
            )

        for batch_id in batch_ids:
            db.execute(
                """
                INSERT OR IGNORE INTO pm_journal_entry_batches (
                    entry_id, batch_id
                )
                VALUES (?, ?)
                """,
                (entry_id, batch_id),
            )

        for item in measurements:
            db.execute(
                """
                INSERT INTO pm_journal_measurements (
                    entry_id, metric, label, value, unit, created_at
                )
                VALUES (?, ?, ?, ?, ?, ?)
                """,
                (
                    entry_id,
                    item["metric"],
                    item["label"],
                    float(item["value"]),
                    item.get("unit"),
                    now,
                ),
            )

        db.commit()
        return entry_id
    finally:
        db.close()


def resolve_follow_up(entry_id, *, user_id=None, user_name=None):
    db = _db()
    try:
        row = db.execute(
            """
            SELECT follow_up_required, resolved_at
            FROM pm_journal_entries
            WHERE id = ?
            """,
            (int(entry_id),),
        ).fetchone()
        if not row:
            raise ValueError("Journal-Eintrag wurde nicht gefunden.")

        if not row["follow_up_required"]:
            raise ValueError("Dieser Eintrag hat keine Folgeaktion.")

        if row["resolved_at"]:
            return False

        db.execute(
            """
            UPDATE pm_journal_entries
            SET resolved_at = ?,
                resolved_by = ?,
                resolved_by_name = ?,
                updated_at = ?
            WHERE id = ?
            """,
            (_now(), user_id, user_name, _now(), int(entry_id)),
        )
        db.commit()
        return True
    finally:
        db.close()


def cancel_journal_entry(
    entry_id,
    *,
    reason,
    user_id=None,
    user_name=None,
):
    reason = _text(reason)
    if len(reason) < 3:
        raise ValueError("Bitte einen Stornogrund angeben.")

    db = _db()
    try:
        row = db.execute(
            "SELECT cancelled_at FROM pm_journal_entries WHERE id = ?",
            (int(entry_id),),
        ).fetchone()
        if not row:
            raise ValueError("Journal-Eintrag wurde nicht gefunden.")

        if row["cancelled_at"]:
            return False

        db.execute(
            """
            UPDATE pm_journal_entries
            SET cancelled_at = ?,
                cancelled_by = ?,
                cancelled_by_name = ?,
                cancel_reason = ?,
                updated_at = ?
            WHERE id = ?
            """,
            (
                _now(),
                user_id,
                user_name,
                reason,
                _now(),
                int(entry_id),
            ),
        )
        db.commit()
        return True
    finally:
        db.close()


def get_revisions(entry_id):
    db = _db()
    try:
        return [
            dict(row)
            for row in db.execute(
                """
                SELECT id, entry_id, revision_no, changed_at,
                       changed_by, changed_by_name
                FROM pm_journal_revisions
                WHERE entry_id = ?
                ORDER BY revision_no DESC
                """,
                (int(entry_id),),
            ).fetchall()
        ]
    finally:
        db.close()


def journal_stats():
    db = _db()
    try:
        today = date.today().isoformat()
        seven_days_ago = (date.today() - timedelta(days=6)).isoformat()

        row = db.execute(
            """
            SELECT
                SUM(
                    CASE
                        WHEN cancelled_at IS NULL
                         AND substr(occurred_at, 1, 10) = ?
                        THEN 1 ELSE 0
                    END
                ) AS today_count,
                SUM(
                    CASE
                        WHEN cancelled_at IS NULL
                         AND substr(occurred_at, 1, 10) >= ?
                        THEN 1 ELSE 0
                    END
                ) AS seven_day_count,
                SUM(
                    CASE
                        WHEN cancelled_at IS NULL
                         AND follow_up_required = 1
                         AND resolved_at IS NULL
                        THEN 1 ELSE 0
                    END
                ) AS open_followups,
                SUM(
                    CASE
                        WHEN cancelled_at IS NULL
                         AND severity = 'critical'
                         AND (
                             follow_up_required = 0
                             OR resolved_at IS NULL
                         )
                        THEN 1 ELSE 0
                    END
                ) AS critical_count
            FROM pm_journal_entries
            """,
            (today, seven_days_ago),
        ).fetchone()

        return {
            "today_count": int(row["today_count"] or 0),
            "seven_day_count": int(row["seven_day_count"] or 0),
            "open_followups": int(row["open_followups"] or 0),
            "critical_count": int(row["critical_count"] or 0),
        }
    finally:
        db.close()


def _magic_valid(extension, content):
    if extension not in ALLOWED_ATTACHMENTS:
        return False

    _, signature = ALLOWED_ATTACHMENTS[extension]

    if extension == ".webp":
        return (
            len(content) >= 12
            and content[:4] == b"RIFF"
            and content[8:12] == b"WEBP"
        )

    return content.startswith(signature)


def save_attachments(
    entry_id,
    uploads,
    *,
    user_id=None,
    user_name=None,
):
    uploads = [item for item in (uploads or []) if item and item.filename]

    if len(uploads) > MAX_ATTACHMENTS_PER_REQUEST:
        raise ValueError(
            f"Maximal {MAX_ATTACHMENTS_PER_REQUEST} Anhänge pro Upload."
        )

    saved = []
    db = _db()

    try:
        for upload in uploads:
            original_name = Path(upload.filename).name
            extension = Path(original_name).suffix.lower()

            if extension not in ALLOWED_ATTACHMENTS:
                raise ValueError(
                    f"Dateityp nicht erlaubt: {original_name}. "
                    "Erlaubt sind JPG, PNG, WEBP und PDF."
                )

            content = upload.stream.read(MAX_ATTACHMENT_BYTES + 1)

            if len(content) > MAX_ATTACHMENT_BYTES:
                raise ValueError(
                    f"Datei zu groß: {original_name}. "
                    "Maximal 8 MB je Datei."
                )

            if not _magic_valid(extension, content):
                raise ValueError(
                    f"Dateiinhalt passt nicht zum Dateityp: {original_name}."
                )

            mime_type = ALLOWED_ATTACHMENTS[extension][0]
            stored_name = f"{uuid.uuid4().hex}{extension}"
            target = UPLOAD_DIR / stored_name
            target.write_bytes(content)

            cur = db.execute(
                """
                INSERT INTO pm_journal_attachments (
                    entry_id, stored_name, original_name,
                    mime_type, size_bytes, uploaded_at,
                    uploaded_by, uploaded_by_name
                )
                VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    int(entry_id),
                    stored_name,
                    original_name,
                    mime_type,
                    len(content),
                    _now(),
                    user_id,
                    user_name,
                ),
            )

            saved.append(int(cur.lastrowid))

        db.commit()
        return saved
    except Exception:
        db.rollback()
        for attachment_id in saved:
            row = db.execute(
                """
                SELECT stored_name
                FROM pm_journal_attachments
                WHERE id = ?
                """,
                (attachment_id,),
            ).fetchone()
            if row:
                try:
                    (UPLOAD_DIR / row["stored_name"]).unlink(missing_ok=True)
                except Exception:
                    pass
        raise
    finally:
        db.close()


def get_attachment(attachment_id):
    db = _db()
    try:
        row = db.execute(
            """
            SELECT *
            FROM pm_journal_attachments
            WHERE id = ? AND deleted_at IS NULL
            """,
            (int(attachment_id),),
        ).fetchone()
        if not row:
            return None

        item = dict(row)
        item["path"] = str(UPLOAD_DIR / item["stored_name"])
        return item
    finally:
        db.close()


def remove_attachment(
    attachment_id,
    *,
    user_id=None,
    user_name=None,
):
    db = _db()
    try:
        row = db.execute(
            """
            SELECT id
            FROM pm_journal_attachments
            WHERE id = ? AND deleted_at IS NULL
            """,
            (int(attachment_id),),
        ).fetchone()
        if not row:
            return False

        db.execute(
            """
            UPDATE pm_journal_attachments
            SET deleted_at = ?,
                deleted_by = ?,
                deleted_by_name = ?
            WHERE id = ?
            """,
            (_now(), user_id, user_name, int(attachment_id)),
        )
        db.commit()
        return True
    finally:
        db.close()


def export_journal_xlsx(entries):
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError as exc:
        raise RuntimeError(
            "Excel-Unterstützung fehlt. Bitte python3-openpyxl installieren."
        ) from exc

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Betriebsjournal"

    headers = [
        "ID",
        "Zeitpunkt",
        "Kategorie",
        "Priorität",
        "Titel",
        "Beschreibung",
        "Pflanzen",
        "Durchgänge",
        "Messwerte",
        "Tags",
        "Folgeaktion",
        "Fällig",
        "Erledigt",
        "Autor",
        "Storniert",
    ]

    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    def safe_text(value):
        if value is None:
            return ""
        text = str(value)
        if text and text[0] in ("=", "+", "-", "@"):
            return "'" + text
        return text

    for row_index, entry in enumerate(entries, start=2):
        measurement_text = ", ".join(
            f"{m['label']}: {m['value']:g} {m['unit'] or ''}".strip()
            for m in entry["measurements"]
        )
        plant_text = ", ".join(
            f"{p['code']} {p['display_name']}"
            for p in entry["plants"]
        )
        batch_text = ", ".join(
            f"{b['code']} {b['name']}"
            for b in entry["batches"]
        )

        values = [
            entry["id"],
            entry["occurred_at"],
            entry["category_label"],
            entry["severity_label"],
            safe_text(entry["title"]),
            safe_text(entry.get("body")),
            safe_text(plant_text),
            safe_text(batch_text),
            safe_text(measurement_text),
            safe_text(entry.get("tags")),
            "Ja" if entry["follow_up_required"] else "Nein",
            entry.get("follow_up_due_on") or "",
            "Ja" if entry.get("resolved_at") else "Nein",
            safe_text(entry.get("created_by_name")),
            "Ja" if entry.get("cancelled_at") else "Nein",
        ]

        for col, value in enumerate(values, start=1):
            ws.cell(row=row_index, column=col, value=value)

    widths = [8, 19, 22, 14, 32, 55, 36, 30, 42, 24, 14, 14, 12, 22, 12]
    for index, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(index)].width = width

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    return stream
