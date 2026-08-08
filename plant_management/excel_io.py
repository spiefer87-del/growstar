from io import BytesIO

from .database import list_cultivars, upsert_cultivar_by_code


HEADERS = [
    ("code", "Code"),
    ("name", "Sorte"),
    ("breeder", "Breeder / Züchter"),
    ("genetics", "Genetik"),
    ("growth_type", "Typ"),
    ("sativa_pct", "Sativa %"),
    ("indica_pct", "Indica %"),
    ("expected_veg_days", "Veg Tage Plan"),
    ("expected_flower_days", "Blüte Tage Plan"),
    ("description", "Beschreibung"),
    ("tags", "Tags"),
    ("notes", "Notizen"),
    ("active", "Aktiv"),
]

HEADER_ALIASES = {
    "code": "code",
    "sorte": "name",
    "name": "name",
    "breeder / züchter": "breeder",
    "breeder": "breeder",
    "züchter": "breeder",
    "genetik": "genetics",
    "genetics": "genetics",
    "typ": "growth_type",
    "sativa %": "sativa_pct",
    "sativa": "sativa_pct",
    "indica %": "indica_pct",
    "indica": "indica_pct",
    "veg tage plan": "expected_veg_days",
    "veg tage": "expected_veg_days",
    "blüte tage plan": "expected_flower_days",
    "blüte tage": "expected_flower_days",
    "beschreibung": "description",
    "tags": "tags",
    "notizen": "notes",
    "aktiv": "active",
}


def _require_openpyxl():
    try:
        import openpyxl
        return openpyxl
    except ImportError as exc:
        raise RuntimeError(
            "Excel-Unterstützung fehlt. Auf Debian/Raspberry Pi bitte "
            "'sudo apt install python3-openpyxl' installieren."
        ) from exc


def _excel_text(value):
    if value is None:
        return ""
    text = str(value)
    # Schutz vor Formula Injection in Exportdateien.
    if text and text[0] in ("=", "+", "-", "@"):
        return "'" + text
    return text


def export_cultivars_xlsx():
    openpyxl = _require_openpyxl()
    from openpyxl.styles import Font, Alignment
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sortenstamm"

    for col, (_, label) in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col, value=label)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    for row_index, cultivar in enumerate(
        list_cultivars(include_inactive=True),
        start=2,
    ):
        for col, (field, _) in enumerate(HEADERS, start=1):
            value = cultivar.get(field)
            if field == "active":
                value = "Ja" if value else "Nein"
            elif isinstance(value, str):
                value = _excel_text(value)
            ws.cell(row=row_index, column=col, value=value)

    widths = [14, 26, 22, 28, 16, 12, 12, 16, 18, 42, 24, 42, 10]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    info = wb.create_sheet("Hinweise")
    info["A1"] = "Growstar Sortenstamm – Excel Import/Export"
    info["A1"].font = Font(bold=True)
    info["A3"] = "Regel"
    info["B3"] = "Beschreibung"
    info["A4"] = "Code"
    info["B4"] = "Bestehende Codes werden beim Import aktualisiert; leere Codes erzeugen neue Datensätze."
    info["A5"] = "Pflichtfeld"
    info["B5"] = "Sorte"
    info["A6"] = "Aktiv"
    info["B6"] = "Ja/Nein, 1/0 oder true/false"
    info.column_dimensions["A"].width = 22
    info.column_dimensions["B"].width = 90

    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    return stream


def template_xlsx():
    openpyxl = _require_openpyxl()
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sortenstamm"

    for col, (_, label) in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col, value=label)
        cell.font = Font(bold=True)

    widths = [14, 26, 22, 28, 16, 12, 12, 16, 18, 42, 24, 42, 10]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width

    ws.freeze_panes = "A2"

    info = wb.create_sheet("Hinweise")
    info.append(["Growstar Sortenstamm Importvorlage"])
    info.append([])
    info.append(["Pflichtfeld", "Sorte"])
    info.append(["Code", "Optional. Bei leerem Code erzeugt Growstar automatisch einen Code."])
    info.append(["Update", "Existiert ein Code bereits, wird dieser Datensatz aktualisiert."])
    info.append(["Aktiv", "Ja/Nein, 1/0 oder true/false"])
    info.column_dimensions["A"].width = 24
    info.column_dimensions["B"].width = 90

    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    return stream


def import_cultivars_xlsx(file_stream):
    openpyxl = _require_openpyxl()

    wb = openpyxl.load_workbook(
        file_stream,
        read_only=True,
        data_only=True,
    )

    ws = wb["Sortenstamm"] if "Sortenstamm" in wb.sheetnames else wb.active

    rows = ws.iter_rows(values_only=True)
    try:
        header_row = next(rows)
    except StopIteration:
        raise ValueError("Die Excel-Datei ist leer.")

    mapping = {}
    for index, header in enumerate(header_row):
        if header is None:
            continue
        normalized = str(header).strip().lower()
        field = HEADER_ALIASES.get(normalized)
        if field:
            mapping[index] = field

    if "name" not in mapping.values():
        raise ValueError("Die Spalte 'Sorte' fehlt.")

    created = 0
    updated = 0
    skipped = 0
    errors = []

    for excel_row, values in enumerate(rows, start=2):
        data = {}

        for index, field in mapping.items():
            value = values[index] if index < len(values) else None
            data[field] = value

        if not any(v not in (None, "") for v in data.values()):
            continue

        if not str(data.get("name") or "").strip():
            skipped += 1
            errors.append(f"Zeile {excel_row}: Sorte fehlt.")
            continue

        active = data.get("active")
        if active is None or active == "":
            data["active"] = True

        try:
            _, action = upsert_cultivar_by_code(data)
            if action == "created":
                created += 1
            else:
                updated += 1
        except Exception as exc:
            skipped += 1
            errors.append(f"Zeile {excel_row}: {exc}")

    return {
        "created": created,
        "updated": updated,
        "skipped": skipped,
        "errors": errors[:50],
    }
