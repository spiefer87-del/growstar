from io import BytesIO

from .database import get_cultivar_by_code
from .propagation import (
    get_seed_lot_by_code,
    list_seed_lots,
    save_seed_lot,
    book_seed_movement,
)


HEADERS = [
    ("code", "Code"),
    ("cultivar_code", "Sorten-Code"),
    ("supplier", "Lieferant"),
    ("breeder_lot", "Hersteller-Lot"),
    ("origin_type", "Herkunft"),
    ("acquired_on", "Eingang"),
    ("produced_on", "Produktionsdatum"),
    ("seed_type", "Samenart"),
    ("storage_location", "Lagerort"),
    ("status", "Status"),
    ("quantity_booking", "Bestandsbuchung"),
    ("notes", "Notizen"),
]


def _require_openpyxl():
    try:
        import openpyxl
        return openpyxl
    except ImportError as exc:
        raise RuntimeError(
            "Excel-Unterstützung fehlt. Bitte python3-openpyxl installieren."
        ) from exc


def _safe_text(value):
    if value is None:
        return ""
    text = str(value)
    if text and text[0] in ("=", "+", "-", "@"):
        return "'" + text
    return text


def export_seed_lots_xlsx():
    openpyxl = _require_openpyxl()
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Saatgut"

    export_headers = [
        "Code",
        "Sorten-Code",
        "Sorte",
        "Lieferant",
        "Hersteller-Lot",
        "Herkunft",
        "Eingang",
        "Produktionsdatum",
        "Samenart",
        "Lagerort",
        "Status",
        "Bestand",
        "Keim-/Erfolgsquote %",
        "Verwendete Einheiten",
        "Notizen",
    ]

    for col, label in enumerate(export_headers, start=1):
        cell = ws.cell(row=1, column=col, value=label)
        cell.font = Font(bold=True)

    for row_index, lot in enumerate(list_seed_lots(), start=2):
        values = [
            lot["code"],
            lot["cultivar_code"],
            lot["cultivar_name"],
            _safe_text(lot.get("supplier")),
            _safe_text(lot.get("breeder_lot")),
            _safe_text(lot.get("origin_type")),
            lot.get("acquired_on") or "",
            lot.get("produced_on") or "",
            _safe_text(lot.get("seed_type")),
            _safe_text(lot.get("storage_location")),
            lot.get("status") or "",
            lot.get("stock", 0),
            lot.get("success_rate"),
            lot.get("used_units", 0),
            _safe_text(lot.get("notes")),
        ]
        for col, value in enumerate(values, start=1):
            ws.cell(row=row_index, column=col, value=value)

    widths = [18, 14, 28, 24, 20, 18, 14, 18, 15, 22, 14, 12, 18, 18, 42]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    return stream


def seed_import_template_xlsx():
    openpyxl = _require_openpyxl()
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Saatgut"

    for col, (_, label) in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col, value=label)
        cell.font = Font(bold=True)

    widths = [18, 14, 24, 20, 18, 14, 18, 15, 22, 14, 18, 42]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width

    ws.freeze_panes = "A2"

    info = wb.create_sheet("Hinweise")
    rows = [
        ["Growstar Saatgut-Import"],
        [],
        ["Sorten-Code", "Pflichtfeld, muss im Sortenstamm existieren."],
        ["Code", "Optional. Bei leerem Code erzeugt Growstar einen Lot-Code."],
        ["Bestandsbuchung", "Bei neuen Lots = Anfangsbestand. Bei bestehenden Lots = zusätzliche +/- Buchung."],
        ["Wichtig", "Bestände werden niemals direkt überschrieben, sondern immer als Bewegung gebucht."],
    ]
    for row in rows:
        info.append(row)
    info.column_dimensions["A"].width = 24
    info.column_dimensions["B"].width = 100

    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    return stream


def import_seed_lots_xlsx(file_stream, *, user_id=None, user_name=None):
    openpyxl = _require_openpyxl()
    wb = openpyxl.load_workbook(
        file_stream,
        read_only=True,
        data_only=True,
    )
    ws = wb["Saatgut"] if "Saatgut" in wb.sheetnames else wb.active

    rows = ws.iter_rows(values_only=True)
    try:
        header = next(rows)
    except StopIteration:
        raise ValueError("Die Excel-Datei ist leer.")

    aliases = {
        "code": "code",
        "sorten-code": "cultivar_code",
        "sortencode": "cultivar_code",
        "lieferant": "supplier",
        "hersteller-lot": "breeder_lot",
        "hersteller lot": "breeder_lot",
        "herkunft": "origin_type",
        "eingang": "acquired_on",
        "produktionsdatum": "produced_on",
        "samenart": "seed_type",
        "lagerort": "storage_location",
        "status": "status",
        "bestandsbuchung": "quantity_booking",
        "notizen": "notes",
    }

    mapping = {}
    for index, value in enumerate(header):
        if value is None:
            continue
        key = str(value).strip().lower()
        if key in aliases:
            mapping[index] = aliases[key]

    if "cultivar_code" not in mapping.values():
        raise ValueError("Die Spalte 'Sorten-Code' fehlt.")

    created = 0
    updated = 0
    booked = 0
    skipped = 0
    errors = []

    for excel_row, values in enumerate(rows, start=2):
        data = {}
        for index, field in mapping.items():
            data[field] = values[index] if index < len(values) else None

        if not any(v not in (None, "") for v in data.values()):
            continue

        cultivar_code = str(data.get("cultivar_code") or "").strip()
        cultivar = get_cultivar_by_code(cultivar_code)
        if not cultivar:
            skipped += 1
            errors.append(
                f"Zeile {excel_row}: Sorten-Code '{cultivar_code}' nicht gefunden."
            )
            continue

        lot_data = {
            "code": data.get("code"),
            "cultivar_id": cultivar["id"],
            "supplier": data.get("supplier"),
            "breeder_lot": data.get("breeder_lot"),
            "origin_type": data.get("origin_type"),
            "acquired_on": data.get("acquired_on"),
            "produced_on": data.get("produced_on"),
            "seed_type": data.get("seed_type"),
            "storage_location": data.get("storage_location"),
            "status": data.get("status") or "available",
            "notes": data.get("notes"),
        }

        booking = data.get("quantity_booking")
        try:
            booking = int(float(booking)) if booking not in (None, "") else 0
        except Exception:
            skipped += 1
            errors.append(
                f"Zeile {excel_row}: Bestandsbuchung ist keine ganze Zahl."
            )
            continue

        try:
            existing = (
                get_seed_lot_by_code(data.get("code"))
                if data.get("code")
                else None
            )

            if existing:
                lot_id = save_seed_lot(
                    lot_data,
                    existing["id"],
                    user_id=user_id,
                    user_name=user_name,
                )
                updated += 1

                if booking:
                    book_seed_movement(
                        lot_id,
                        booking,
                        "adjustment",
                        note=f"Excel-Import Zeile {excel_row}",
                        user_id=user_id,
                        user_name=user_name,
                    )
                    booked += 1
            else:
                if booking < 0:
                    raise ValueError(
                        "Ein neues Saatgut-Lot kann nicht mit negativem "
                        "Anfangsbestand angelegt werden."
                    )

                lot_id = save_seed_lot(
                    lot_data,
                    initial_quantity=booking,
                    user_id=user_id,
                    user_name=user_name,
                )
                created += 1
                if booking > 0:
                    booked += 1

        except Exception as exc:
            skipped += 1
            errors.append(f"Zeile {excel_row}: {exc}")

    return {
        "created": created,
        "updated": updated,
        "booked": booked,
        "skipped": skipped,
        "errors": errors[:60],
    }
